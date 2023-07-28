"""
In order to perform actions on Excel File, we need a library...
Here we are using --> openpyxl

Command --> pip install openpyxl

To Confirm --> File > Settings > Project: <Project_name> > Python Interpreter >
                    Click on plus > Search it in Available Packages...
"""
import os

import openpyxl

# Declaration of absolute path...
my_path = os.path.abspath(os.path.dirname(__file__))
path = os.path.join(my_path, "DataFiles\\Employee.xlsx")

# Declaration of Workbook & Sheet...
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]

# Getting Rows and Columns...
rows = int(sheet.max_row)
columns = int(sheet.max_column)
print("Row : ", rows)
print("Column :", columns)

# Looping the data...
for r in range(1, rows + 1):
    for c in range(1, columns + 1):
        print(sheet.cell(r, c).value, end="  |  ")
    print()
